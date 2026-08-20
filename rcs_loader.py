"""
Input loader for Karix RCS Bot Builder & DLT templates.

Turns raw rows (CSV, XLSX, JSON) into validated RcsTemplateSubmission objects.
Supports text messages, rich cards, suggestions (URL, Reply, Dialer), and variables.
"""

import csv
import json
import logging
import os
import re
from pathlib import Path
from rcs_config import get_rcs_bot_id, get_rcs_entity_id
from rcs_models import RcsTemplateSubmission

logger = logging.getLogger(__name__)


def infer_cta_link_and_button(text: str) -> tuple[str, str]:
    """
    If no explicit CTA button or link is given in the brief:
    Infers the most relevant Tata Capital landing page and button label based on copy keywords.
    """
    t_lower = text.lower()

    if any(w in t_lower for w in ("home loan", "housing", "property", "mortgage")):
        return "Check Rates", "https://www.tatacapital.com/home-loan.html"
    elif any(w in t_lower for w in ("business loan", "enterprise", "msme", "sme", "working capital")):
        return "Apply Business", "https://www.tatacapital.com/business-loan.html"
    elif any(w in t_lower for w in ("vehicle", "car", "2-wheeler", "bike", "auto")):
        return "Explore Vehicle Loan", "https://www.tatacapital.com/vehicle-loan.html"
    elif any(w in t_lower for w in ("eligibility", "eligible", "check offer", "check my offer")):
        return "Check Eligibility", "https://www.tatacapital.com/personal-loan.html"
    elif any(w in t_lower for w in ("claim", "pre-approved", "pre approved", "exclusive")):
        return "Claim Your Offer", "https://www.tatacapital.com/personal-loan.html"
    elif any(w in t_lower for w in ("feedback", "survey", "rating", "experience", "satisfied")):
        return "Rate Experience", "https://www.tatacapital.com"
    else:
        return "Apply Now", "https://www.tatacapital.com/personal-loan.html"


def parse_single_cell_card_block(cell_text: str) -> dict:
    """
    Decompose an unstructured marketing card text block (from Excel cells) into:
    - card_title
    - card_description
    - button_text
    - button_url
    """
    if not cell_text:
        return {
            "card_title": "",
            "card_description": "",
            "button_text": "Apply Now",
            "button_url": "https://www.tatacapital.com/personal-loan.html",
        }

    lines = [line.strip() for line in cell_text.splitlines() if line.strip()]

    inferred_text, inferred_url = infer_cta_link_and_button(cell_text)
    button_text = inferred_text
    button_url = inferred_url
    clean_lines = []

    for line in lines:
        is_cta = False
        # 1. Match CTA Button <Text> or CTA button<Text> or CTA<Text>
        m_cta = re.search(r'CTA\s*(?:Button|button)?\s*[:<\[]\s*([^>\]<]+)\s*[>\]]', line, re.IGNORECASE)
        if m_cta:
            cand = m_cta.group(1).strip()
            if cand.lower() not in ("link", "url"):
                button_text = cand
            is_cta = True

        # 2. Match [Button Text] <link> or <Button Text> <link>
        m_link = re.search(r'[\[<]([^>\]<]+)[\]>]\s*(?:<link>|\[link\])', line, re.IGNORECASE)
        if m_link and not is_cta:
            cand = m_link.group(1).strip()
            if cand.lower() not in ("link", "url"):
                button_text = cand
            is_cta = True

        # 3. Match prompt lines like "Tap to proceed ⬇️" or "Tap below to check eligibility⬇️"
        if re.search(r'^(?:Tap|Click|Press)\s+(?:below|here|to\s+proceed|to\s+check|to\s+apply).*?(?:⬇️|->|:|here)?$', line, re.IGNORECASE):
            is_cta = True

        if not is_cta:
            clean_lines.append(line)

    full_desc = "\n\n".join(clean_lines)
    first_line = clean_lines[0] if clean_lines else "Special Offer"
    clean_title = re.sub(r'<[^>]+>|\[[^\]]+\]|\{[^}]+\}', '', first_line).strip()
    clean_title = re.sub(r'^[,\s:–—\-]+|[,\s:–—\-]+$', '', clean_title)
    if re.match(r'^(?:Dear|Hi|Hello)\b', clean_title, re.IGNORECASE) or len(clean_title) < 4:
        clean_title = "Pre-Approved Loan Offer ✨"

    return {
        "card_title": clean_title[:100],
        "card_description": full_desc,
        "button_text": button_text,
        "button_url": button_url,
    }

def _build_suggestions_from_row(row: dict) -> list[dict]:
    """Parse button columns into Karix RCS suggestion dictionaries."""
    suggestions = []

    # If suggestions array is already provided as JSON
    if row.get("suggestions"):
        if isinstance(row["suggestions"], list):
            return row["suggestions"]
        try:
            return json.loads(row["suggestions"])
        except (json.JSONDecodeError, TypeError):
            pass

    btype = (row.get("button_type") or row.get("suggestion_type") or "").strip().upper()
    btext = (row.get("button_text") or row.get("suggestion_text") or "").strip()
    burl = (row.get("button_url") or row.get("url") or "").strip()
    bphone = (row.get("button_phone") or row.get("phone") or row.get("phone_number") or "").strip()

    if btext:
        if "|" in btext and btype in ("", "REPLY", "SUGGESTION"):
            for item in btext.split("|"):
                clean = item.strip()
                if clean:
                    suggestions.append({
                        "suggestionType": "reply",
                        "text": clean,
                        "postbackData": clean.lower().replace(" ", "_"),
                    })
        elif btype in ("DIALER", "DIALER_ACTION", "CALL", "PHONE") or bphone:
            suggestions.append({
                "suggestionType": "dialer_action",
                "text": btext or "Call Now",
                "postbackData": btext.lower().replace(" ", "_") if btext else "call_now",
                "phoneNumber": bphone or "+919999999999",
            })
        elif btype in ("URL", "URL_ACTION", "LINK") or burl:
            suggestions.append({
                "suggestionType": "url_action",
                "text": btext or "Apply Now",
                "postbackData": btext.lower().replace(" ", "_") if btext else "apply_now",
                "url": burl or "https://www.tatacapital.com",
            })
        else:
            suggestions.append({
                "suggestionType": "reply",
                "text": btext,
                "postbackData": btext.lower().replace(" ", "_"),
            })

    return suggestions


def _normalize_row_keys(row: dict) -> dict:
    """Normalize spreadsheet header keys: strip whitespace, map common aliases to canonical names."""
    aliases = {
        "templatename": "template_name",
        "temlatename": "template_name",
        "templte_name": "template_name",
        "campaignname": "campaign_name",
        "botid": "bot_id",
        "senderid": "sender_id",
        "templatetype": "template_type",
        "mediaurl": "media_url",
        "imageurl": "image_url",
        "image": "image",
        "cardtitle": "card_title",
        "carddescription": "card_description",
        "textmessage": "text_message",
        "templatemessage": "template_message",
        "entityid": "entity_id",
        "sourceref": "source_ref",
        "buttontext": "button_text",
        "buttonurl": "button_url",
        "buttonphone": "button_phone",
        "buttontype": "button_type",
        "suggestiontype": "suggestion_type",
        "suggestiontext": "suggestion_text",
    }
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        canonical = str(key).strip()
        compact = re.sub(r"[^a-z0-9]", "", canonical.lower())
        normalized[aliases.get(compact, canonical)] = value
    return normalized


def _parse_sender_ids(raw) -> list[str]:
    """Parse sender IDs from a pipe/comma-separated string, list, or None."""
    if not raw:
        return []
    if isinstance(raw, (list, tuple)):
        return [str(x).strip() for x in raw if str(x).strip()]
    return [x.strip() for x in re.split(r"[|,]", str(raw)) if x.strip()]




def _build_carousel_cards_from_row(row: dict) -> list[dict]:
    """Parse multiple cards for carousel templates from pipe-separated columns or JSON."""
    if row.get("carousel_cards"):
        if isinstance(row["carousel_cards"], list):
            return row["carousel_cards"]
        try:
            return json.loads(row["carousel_cards"])
        except (json.JSONDecodeError, TypeError):
            pass

    titles = [t.strip() for t in str(row.get("card_title") or row.get("title") or "").split("|") if t.strip()]
    descriptions = [d.strip() for d in str(row.get("body") or row.get("card_description") or row.get("text_message") or row.get("description") or "").split("|") if d.strip()]
    media_urls = [u.strip() for u in str(row.get("media_url") or row.get("image_url") or row.get("image") or "").split("|") if u.strip()]
    button_texts = [b.strip() for b in str(row.get("button_text") or row.get("button_name") or "").split("|") if b.strip()]
    button_urls = [u.strip() for u in str(row.get("button_url") or row.get("link") or "").split("|") if u.strip()]
    button_types = [t.strip().upper() for t in str(row.get("button_type") or "").split("|") if t.strip()]

    max_cards = max(len(titles), len(descriptions), len(media_urls), len(button_texts), 2)

    cards = []
    for i in range(max_cards):
        c_title = titles[i] if i < len(titles) else (f"Card {i+1}" if titles else "")
        c_desc = descriptions[i] if i < len(descriptions) else (descriptions[0] if descriptions else "")
        c_url = media_urls[i] if i < len(media_urls) else (media_urls[0] if media_urls else "https://www.tatacapital.com/content/dam/tata-capital/header-logo/tata-capital-logo.png")

        card_suggs = []
        if i < len(button_texts):
            btext = button_texts[i]
            btype = button_types[i] if i < len(button_types) else (button_types[0] if button_types else "URL")
            b_link = button_urls[i] if i < len(button_urls) else (button_urls[0] if button_urls else "https://www.tatacapital.com")

            if btype in ("URL", "URL_ACTION", "LINK") or b_link:
                card_suggs.append({
                    "suggestionType": "url_action",
                    "text": btext,
                    "postbackData": btext.lower().replace(" ", "_"),
                    "url": b_link,
                })
            else:
                card_suggs.append({
                    "suggestionType": "reply",
                    "text": btext,
                    "postbackData": btext.lower().replace(" ", "_"),
                })

        cards.append({
            "cardTitle": c_title,
            "cardDescription": c_desc,
            "mediaUrl": c_url,
            "suggestions": card_suggs,
        })

    return cards


def _row_to_rcs_submission(row: dict, client: str = "tata", fallback_idx: int = 1) -> RcsTemplateSubmission:
    """Convert a normalized dict to an RcsTemplateSubmission."""
    row = _normalize_row_keys(row)
    c = (row.get("client") or client).lower()

    template_name = str(
        row.get("template_name")
        or row.get("templatename")
        or row.get("name")
        or row.get("template")
        or row.get("campaign_name")
        or row.get("campaign")
        or f"rcs_template_{fallback_idx}"
    ).strip()

    bot_id = str(row.get("bot_id") or row.get("sender_id") or get_rcs_bot_id(c)).strip()

    raw_type = str(row.get("template_type") or row.get("type") or "").strip().lower()
    header_type = str(row.get("header_type") or "").strip().lower()
    media_url = str(row.get("media_url") or row.get("image_url") or row.get("image") or "").strip() or None
    card_title = str(row.get("card_title") or row.get("title") or row.get("header_text") or "").strip() or None

    is_carousel = (
        raw_type in ("carousel", "carousal", "carousel_cards", "multi_card")
        or bool(row.get("carousel_cards"))
        or ("|" in str(row.get("card_title") or "") and "|" in str(row.get("media_url") or ""))
        or ("|" in str(row.get("card_title") or "") and raw_type in ("carousel", "carousal"))
    )

    public_base = os.environ.get("RENDER_EXTERNAL_URL") or os.environ.get("PUBLIC_APP_URL") or "https://whitelisting-agent.onrender.com"

    # Resolve default media fallback per official spec ratio
    orientation_key = str(row.get("orientation") or "VERTICAL").strip().upper()
    height_key = str(row.get("height") or "MEDIUM").strip().upper()
    if orientation_key == "HORIZONTAL":
        _fallback_media = f"{public_base}/api/media/default_rcs_3x4.png"
    elif height_key == "SHORT":
        _fallback_media = f"{public_base}/api/media/default_rcs_3x1.png"
    else:
        _fallback_media = f"{public_base}/api/media/default_rcs_2x1.png"

    if is_carousel:
        template_type = "carousel"
        carousel_cards = _build_carousel_cards_from_row(row)
    elif raw_type in ("richcard", "card", "image") or header_type in ("image", "media", "richcard") or media_url or card_title:
        template_type = "richcard"
        carousel_cards = []
        if not media_url:
            media_url = _fallback_media
    else:
        template_type = "text"
        carousel_cards = []
    message = (
        row.get("text_message")
        or row.get("body")
        or row.get("card_description")
        or row.get("template_message")
        or row.get("description")
        or row.get("message")
        or ""
    ).strip()

    suggestions = _build_suggestions_from_row(row)
    category = str(row.get("category") or row.get("template_category") or "TRANSACTIONAL").strip().upper()

    return RcsTemplateSubmission(
        template_name=template_name,
        bot_id=bot_id,
        template_type=template_type,
        text_message=message,
        card_title=card_title,
        card_description=message if template_type == "richcard" else None,
        media_url=media_url,
        orientation=str(row.get("orientation") or "VERTICAL").strip().upper(),
        height=str(row.get("height") or "MEDIUM").strip().upper(),
        width=str(row.get("width") or "MEDIUM").strip().upper(),
        suggestions=suggestions,
        carousel_cards=carousel_cards,
        template_category=category,
        entity_id=str(row.get("entity_id") or get_rcs_entity_id(c)).strip(),
        client=c,
        channel="rcs",
        source_ref=row.get("source_ref") or template_name,
    )


def load_rcs_from_csv(path: str, client: str = "tata") -> list[RcsTemplateSubmission]:
    """Load RCS templates from a CSV file."""
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for idx, raw_row in enumerate(csv.DictReader(f), 1):
            if not any(raw_row.values()):
                continue
            rows.append(_row_to_rcs_submission(raw_row, client=client, fallback_idx=idx))
    return rows

# ---------------------------------------------------------------------------
# Official Karix RCS media specifications (RCS specifications docx)
#   Rich Card: VERTICAL SHORT=3:1(1440x480)  VERTICAL MEDIUM=2:1(1440x720)  HORIZONTAL=3:4(768x1024)  max 2MB
#   Carousel : SHORT SMALL=8:5(1160x720)  SHORT MEDIUM=5:2(1800x720)  MEDIUM SMALL=1:1(770x720)  MEDIUM MEDIUM=16:9(1280x720)  max 1MB
# ---------------------------------------------------------------------------

RICH_CARD_IMAGE_SPECS: dict[tuple, dict] = {
    ("VERTICAL", "SHORT"): {"ratio": (3, 1), "optimal": (1440, 480), "max_bytes": 2 * 1024 * 1024},
    ("VERTICAL", "MEDIUM"): {"ratio": (2, 1), "optimal": (1440, 720), "max_bytes": 2 * 1024 * 1024},
    ("HORIZONTAL", "SHORT"): {"ratio": (3, 4), "optimal": (768, 1024), "max_bytes": 2 * 1024 * 1024},
    ("HORIZONTAL", "MEDIUM"): {"ratio": (3, 4), "optimal": (768, 1024), "max_bytes": 2 * 1024 * 1024},
}

CAROUSEL_IMAGE_SPECS: dict[tuple, dict] = {
    ("SHORT", "SMALL"): {"ratio": (8, 5), "optimal": (1160, 720), "max_bytes": 1 * 1024 * 1024},
    ("SHORT", "MEDIUM"): {"ratio": (5, 2), "optimal": (1800, 720), "max_bytes": 1 * 1024 * 1024},
    ("MEDIUM", "SMALL"): {"ratio": (1, 1), "optimal": (770, 720), "max_bytes": 1 * 1024 * 1024},
    ("MEDIUM", "MEDIUM"): {"ratio": (16, 9), "optimal": (1280, 720), "max_bytes": 1 * 1024 * 1024},
}

ACCEPTED_IMAGE_FORMATS = (".jpg", ".jpeg", ".png", ".gif")
ACCEPTED_VIDEO_FORMATS = (".mp4", ".m4v", ".mpeg", ".webm", ".h263", ".m4p")

def _spec_for_richcard(sub: RcsTemplateSubmission) -> dict:
    orientation = (getattr(sub, "orientation", "VERTICAL") or "VERTICAL").upper()
    height = (getattr(sub, "height", "MEDIUM") or "MEDIUM").upper()
    return RICH_CARD_IMAGE_SPECS.get(
        (orientation, height),
        RICH_CARD_IMAGE_SPECS[("VERTICAL", "MEDIUM")],
    )

def _spec_for_carousel(sub: RcsTemplateSubmission) -> dict:
    height = (getattr(sub, "height", "MEDIUM") or "MEDIUM").upper()
    width = (getattr(sub, "width", "MEDIUM") or "MEDIUM").upper()
    return CAROUSEL_IMAGE_SPECS.get(
        (height, width),
        CAROUSEL_IMAGE_SPECS[("MEDIUM", "MEDIUM")],
    )

def _fit_rcs_image(img_bytes: bytes, spec: dict) -> bytes:
    """
    Fit image to the official spec ratio, resize to optimal resolution, and compress
    below the max file size using progressive JPEG quality reduction.
    """
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            bg.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")

        current_w, current_h = img.size
        ratio_w, ratio_h = spec["ratio"]
        target_aspect = ratio_w / ratio_h
        current_aspect = current_w / current_h

        # Center-crop to the required aspect ratio
        if abs(current_aspect - target_aspect) > 0.02:
            if current_aspect > target_aspect:
                new_w = int(current_h * target_aspect)
                left = (current_w - new_w) // 2
                img = img.crop((left, 0, left + new_w, current_h))
            else:
                new_h = int(current_w / target_aspect)
                top = (current_h - new_h) // 2
                img = img.crop((0, top, current_w, top + new_h))

        # Resize down to optimal resolution if larger
        opt_w, opt_h = spec["optimal"]
        if img.width > opt_w or img.height > opt_h:
            img.thumbnail((opt_w, opt_h), Image.Resampling.LANCZOS)

        max_bytes = spec["max_bytes"]
        # Progressive quality reduction until under max file size
        quality = 92
        while quality >= 45:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            data = buf.getvalue()
            if len(data) <= max_bytes:
                return data
            quality -= 10
        return data
    except Exception:
        return img_bytes

def _ensure_aspect_ratio(img_bytes: bytes, target_ratio: tuple = (16, 9)) -> bytes:
    """
    Auto-fit image bytes to the REQUESTED target ratio (e.g. 2:1 for standalone rich cards,
    3:4 for carousel cards). Preserves the image only if it already matches the target.
    """
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes))
        current_w, current_h = img.size
        current_aspect = current_w / current_h

        target_w, target_h = target_ratio
        target_aspect = target_w / target_h

        # Only preserve if the image already matches the requested target ratio
        if abs(current_aspect - target_aspect) < 0.08:
            return img_bytes

        # Crop to the requested target ratio (center crop)
        if current_aspect > target_aspect:
            new_w = int(current_h * target_aspect)
            left = (current_w - new_w) // 2
            img = img.crop((left, 0, left + new_w, current_h))
        else:
            new_h = int(current_w / target_aspect)
            top = (current_h - new_h) // 2
            img = img.crop((0, top, current_w, top + new_h))

        buf = io.BytesIO()
        fmt = "PNG" if img.format == "PNG" else "JPEG"
        img.save(buf, format=fmt, quality=95)
        return buf.getvalue()
    except Exception:
        return img_bytes


def _extract_images_from_xlsx(path: str) -> list[tuple[str, bytes]]:
    """Extract embedded media (images/videos) from an Excel (.xlsx) file in order, preserving original bytes."""
    media = []
    try:
        import zipfile
        with zipfile.ZipFile(path, "r") as z:
            media_names = [f for f in z.namelist() if f.startswith("xl/media/")]
            def natural_key(name):
                return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', name)]
            media_names.sort(key=natural_key)
            for name in media_names:
                filename = name.split("/")[-1]
                media.append((filename, z.read(name)))
    except Exception as e:
        logger.warning("Could not extract embedded media from xlsx: %s", e)
    return media

def _upload_and_bind_rcs_images(raw_media: list[tuple[str, bytes]], subs: list[RcsTemplateSubmission], client: str) -> None:
    """
    Fit each extracted image to the official RCS spec ratio for the template's
    orientation/height/width, compress to the max file size, upload, and bind the
    Karix fileName back onto the templates.
    """
    if not raw_media:
        return
    try:
        from rcs_client import upload_rcs_media
    except Exception:
        return
    for sub in subs:
        if sub.template_type == "richcard" and raw_media:
            try:
                fname, media_data = raw_media[0]
                ext = Path(fname).suffix.lower()
                if ext in ACCEPTED_VIDEO_FORMATS:
                    fitted = media_data
                    ratio_label = "video"
                else:
                    spec = _spec_for_richcard(sub)
                    fitted = _fit_rcs_image(media_data, spec)
                    ratio_label = str(spec["ratio"])
                k_name = upload_rcs_media(fitted, filename=fname, client=client)
                setattr(sub, "file_name", k_name)
                logger.info("Bound rich card media %s -> %s (ratio %s)", fname, k_name, ratio_label)
            except Exception as ex:
                logger.warning("Failed to bind rich card media: %s", ex)
        elif sub.template_type == "carousel" and sub.carousel_cards:
            spec = _spec_for_carousel(sub)
            for c_idx, card in enumerate(sub.carousel_cards):
                if c_idx >= len(raw_media):
                    break
                try:
                    fname, media_data = raw_media[c_idx]
                    ext = Path(fname).suffix.lower()
                    if ext in ACCEPTED_VIDEO_FORMATS:
                        fitted = media_data
                    else:
                        fitted = _fit_rcs_image(media_data, spec)
                    k_name = upload_rcs_media(fitted, filename=fname, client=client)
                    card["fileName"] = k_name
                except Exception as ex:
                    logger.warning("Failed to bind carousel card media %s: %s", fname, ex)

def load_rcs_from_excel(path: str, client: str = "tata") -> list[RcsTemplateSubmission]:
    """Load RCS templates from an Excel (.xlsx) file with auto-extracted embedded images."""
    import openpyxl

    # Extract raw embedded media without transforming them
    raw_media = _extract_images_from_xlsx(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    all_raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    first_row = [str(c or "").strip().lower() for c in all_raw_rows[0]] if all_raw_rows else []
    has_standard_headers = any(h in ("template_name", "templatename", "name", "body", "body_text", "components", "category", "language", "card_title", "card_description", "media_url", "header_type") for h in first_row)

    # Check if there is a row that contains multiple rich single-cell card blocks (only when no standard column headers exist)
    block_row_cards = None
    if not has_standard_headers:
        for r in all_raw_rows:
            text_blocks = [str(c).strip() for c in r if c is not None and len(str(c).strip()) > 35]
            if len(text_blocks) >= 2:
                block_row_cards = text_blocks
                break
        if block_row_cards:
            c_cards = []
            for c_idx, block in enumerate(block_row_cards):
                card_dict = parse_single_cell_card_block(block)

                b_text = card_dict.get("button_text") or "Apply Now"
                b_url = card_dict.get("button_url") or "https://www.tatacapital.com"
                card_dict["suggestions"] = [
                    {
                        "suggestionType": "url_action",
                        "text": b_text,
                        "postbackData": b_text.lower().replace(" ", "_"),
                        "url": b_url,
                    }
                ]
                c_cards.append(card_dict)

            base_name = Path(path).stem
            clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', base_name).strip('_')
            t_name = f"tata_{clean_name}_carousel"[:25].rstrip('_')

            sub = RcsTemplateSubmission(
                template_name=t_name,
                bot_id=get_rcs_bot_id(client),
                template_type="carousel",
                carousel_cards=c_cards,
                template_category="TRANSACTIONAL",
                entity_id=get_rcs_entity_id(client),
                client=client.lower(),
                channel="rcs",
                source_ref=t_name,
            )
            _upload_and_bind_rcs_images(raw_media, [sub], client)
            return [sub]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        raw_row = {}
        for h, val in zip(headers, row):
            if h:
                raw_row[h] = str(val).strip() if val is not None else ""

        if not any(raw_row.values()):
            continue

        sub = _row_to_rcs_submission(raw_row, client=client, fallback_idx=idx)
        rows.append(sub)

    _upload_and_bind_rcs_images(raw_media, rows, client)
    return rows


def load_rcs_from_list(rows: list[dict], client: str = "tata") -> list[RcsTemplateSubmission]:
    """Load from a list of dicts already in memory."""
    return [_row_to_rcs_submission(row, client=client, fallback_idx=idx) for idx, row in enumerate(rows, 1)]
