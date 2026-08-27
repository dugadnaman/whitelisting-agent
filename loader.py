"""
Input loader: turns raw rows (CSV, JSON list, or Excel files) into validated
TemplateSubmission objects for WhatsApp.

Supports both standard column spreadsheets and unstructured single-cell
marketing briefs (like Book 4) with automatic image extraction.
"""

import csv
import json
import logging
import re
import zipfile
from pathlib import Path

from config import get_waba_id
from models import TemplateComponent, TemplateSubmission

logger = logging.getLogger(__name__)


def infer_whatsapp_cta(text: str, client: str = "bajaj") -> tuple[str, str, str]:
    """
    If no explicit CTA is provided:
    Infers the appropriate button text, dynamic URL, and Meta sample example.
    """
    t_lower = text.lower()
    is_tata = client.lower() == "tata"

    if is_tata:
        base_domain = "https://www.tatacapital.com"
        if any(w in t_lower for w in ("home loan", "housing", "property")):
            return (
                "Check Rates",
                f"{base_domain}/home-loan.html/{{{{1}}}}",
                f"{base_domain}/home-loan.html",
            )
        elif any(w in t_lower for w in ("business loan", "enterprise", "msme")):
            return (
                "Apply Business",
                f"{base_domain}/business-loan.html/{{{{1}}}}",
                f"{base_domain}/business-loan.html",
            )
        elif any(w in t_lower for w in ("vehicle", "car", "2-wheeler", "bike")):
            return (
                "Explore Vehicle",
                f"{base_domain}/vehicle-loan.html/{{{{1}}}}",
                f"{base_domain}/vehicle-loan.html",
            )
        elif any(w in t_lower for w in ("eligibility", "eligible", "check offer")):
            return (
                "Check Eligibility",
                f"{base_domain}/personal-loan.html/{{{{1}}}}",
                f"{base_domain}/personal-loan.html",
            )
        elif any(w in t_lower for w in ("claim", "pre-approved", "pre approved", "exclusive")):
            return (
                "Claim Your Offer",
                f"{base_domain}/personal-loan.html/{{{{1}}}}",
                f"{base_domain}/personal-loan.html",
            )
        else:
            return (
                "Apply Now",
                f"{base_domain}/personal-loan.html/{{{{1}}}}",
                f"{base_domain}/personal-loan.html",
            )
    else:
        # Bajaj
        base_domain = "https://www.bajajfinservmarkets.in"
        return "Apply Now", "https://1kx.in/{{1}}", base_domain


def parse_single_cell_whatsapp_block(cell_text: str, client: str = "bajaj") -> dict:
    """
    Decompose an unstructured marketing card text block (from Excel cells) into WhatsApp components:
    - Clean body (with CTA instruction lines stripped)
    - Normalized variables ({{1}}, {{2}}...)
    - Clean header title
    - Action button with dynamic link
    """
    if not cell_text:
        return {
            "header": None,
            "body": "",
            "button_text": "Apply Now",
            "button_url": "https://1kx.in/{{1}}",
            "button_example": "https://www.tatacapital.com",
        }

    lines = [line.strip() for line in cell_text.splitlines() if line.strip()]

    inferred_btn, inferred_url, inferred_ex = infer_whatsapp_cta(cell_text, client=client)
    button_text = inferred_btn
    button_url = inferred_url
    button_example = inferred_ex
    clean_lines = []

    for line in lines:
        is_cta = False
        # 1. Match CTA Button <Text> or CTA button<Text> or CTA<Text>
        m_cta = re.search(
            r"CTA\s*(?:Button|button)?\s*[:<\[]\s*([^>\]<]+)\s*[>\]]",
            line,
            re.IGNORECASE,
        )
        if m_cta:
            cand = m_cta.group(1).strip()
            if cand.lower() not in ("link", "url"):
                button_text = cand
            is_cta = True

        # 2. Match [Button Text] <link> or <Button Text> <link>
        m_link = re.search(r"[\[<]([^>\]<]+)[\]>]\s*(?:<link>|\[link\])", line, re.IGNORECASE)
        if m_link and not is_cta:
            cand = m_link.group(1).strip()
            if cand.lower() not in ("link", "url"):
                button_text = cand
            is_cta = True

        # 3. Match prompt lines like "Tap to proceed ⬇️" or "Tap below to check eligibility⬇️"
        if re.search(
            r"^(?:Tap|Click|Press)\s+(?:below|here|to\s+proceed|to\s+check|to\s+apply).*?(?:⬇️|->|:|here)?$",
            line,
            re.IGNORECASE,
        ):
            is_cta = True

        if not is_cta:
            clean_lines.append(line)

    full_desc = "\n\n".join(clean_lines)

    # Normalize body variables: <Name>, <amt>, [var] -> {{1}}, {{2}}
    spaced_body = re.sub(r"([A-Za-z0-9])(<[^>]+>)", r"\1 \2", full_desc)
    spaced_body = re.sub(r"(<[^>]+>)([A-Za-z0-9])", r"\1 \2", spaced_body)
    spaced_body = re.sub(r"([A-Za-z0-9])(\{#[^#]+#\})", r"\1 \2", spaced_body)

    placeholders = []

    def _repl(m):
        idx = len(placeholders) + 1
        placeholders.append(m.group(0))
        return f"{{{{{idx}}}}}"

    pattern = r"(\{\{\d+\}\}|\{\{[a-zA-Z0-9_]+\}\}|<[^>]+>|\{#[^#]+#\}|\[[a-zA-Z0-9_]+\]|\{[a-zA-Z0-9_]+\})"
    normalized_body = re.sub(pattern, _repl, spaced_body)

    # Extract clean header title
    first_line = clean_lines[0] if clean_lines else "Special Offer"
    clean_title = re.sub(r"<[^>]+>|\[[^\]]+\]|\{[^}]+\}", "", first_line).strip()
    clean_title = re.sub(r"^[,\s:–—\-]+|[,\s:–—\-]+$", "", clean_title)
    if re.match(r"^(?:Dear|Hi|Hello)\b", clean_title, re.IGNORECASE) or len(clean_title) < 4:
        clean_title = "Pre-Approved Personal Loan ✨"

    return {
        "header": clean_title[:60],
        "body": normalized_body,
        "button_text": button_text,
        "button_url": button_url,
        "button_example": button_example,
    }


def _detect_media_kind(filename: str, raw_bytes: bytes) -> tuple[str, str]:
    """
    Detect whether raw media bytes are IMAGE, VIDEO, or DOCUMENT, and return (kind, mime_type).
    """
    fn = filename.lower()
    if fn.endswith((".mp4", ".m4v")):
        return "VIDEO", "video/mp4"
    if fn.endswith((".mov", ".qt")):
        return "VIDEO", "video/quicktime"
    if fn.endswith((".avi", ".mkv", ".webm", ".3gp", ".wmv")):
        return "VIDEO", "video/mp4"
    if fn.endswith((".png",)):
        return "IMAGE", "image/png"
    if fn.endswith((".jpg", ".jpeg")):
        return "IMAGE", "image/jpeg"
    if fn.endswith((".webp",)):
        return "IMAGE", "image/webp"
    if fn.endswith((".gif",)):
        return "IMAGE", "image/gif"
    if fn.endswith((".pdf",)):
        return "DOCUMENT", "application/pdf"

    # Check by magic bytes
    if raw_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        return "IMAGE", "image/png"
    if raw_bytes.startswith(b"\xff\xd8\xff"):
        return "IMAGE", "image/jpeg"
    if raw_bytes.startswith(b"%PDF"):
        return "DOCUMENT", "application/pdf"
    if len(raw_bytes) > 12 and (b"ftyp" in raw_bytes[:16] or b"moov" in raw_bytes[:32] or b"mdat" in raw_bytes[:32]):
        return "VIDEO", "video/mp4"

    return "IMAGE", "image/png"


def _extract_media_from_xlsx(path: str) -> list[dict]:
    """
    Extract all embedded media (images, videos, documents) from an Excel (.xlsx) file in order.
    Returns list of dicts: {"filename": str, "bytes": bytes, "kind": str, "mime_type": str}
    """
    media_items = []
    try:
        with zipfile.ZipFile(path, "r") as z:
            media_names = [f for f in z.namelist() if f.startswith(("xl/media/", "xl/embeddings/"))]

            def natural_key(name):
                return [int(c) if c.isdigit() else c for c in re.split(r"(\d+)", name)]

            media_names.sort(key=natural_key)
            for name in media_names:
                filename = name.split("/")[-1]
                raw_data = z.read(name)
                kind, mime_type = _detect_media_kind(filename, raw_data)
                media_items.append(
                    {
                        "filename": filename,
                        "bytes": raw_data,
                        "kind": kind,
                        "mime_type": mime_type,
                    }
                )
    except Exception as e:
        logger.warning("Could not extract embedded media from xlsx: %s", e)
    return media_items


def _extract_images_from_xlsx(path: str) -> list[tuple[str, bytes]]:
    """Legacy image extractor helper."""
    return [(m["filename"], m["bytes"]) for m in _extract_media_from_xlsx(path) if m["kind"] == "IMAGE"]


def _row_to_submission(row: dict, client: str = "bajaj") -> TemplateSubmission:
    _TC_FIELDS = {
        "type",
        "text",
        "format",
        "variables",
        "buttons",
        "example",
        "media_url",
        "media_file",
        "image_bytes",
        "file_type",
    }

    components = []
    for c in row.get("components", []):
        if isinstance(c, dict):
            if c.keys() <= _TC_FIELDS:
                components.append(TemplateComponent(**c))
            else:
                components.append(c)
        else:
            components.append(c)
    c_client = row.get("client") or client
    template_name = row.get("template_name") or row.get("name")
    return TemplateSubmission(
        client=c_client,
        channel=row.get("channel", "whatsapp"),
        template_name=template_name,
        language=row.get("language", "en"),
        category=row.get("category", "MARKETING"),
        waba_id=row.get("waba_id") or _resolve_row_waba(c_client, {}),
        components=components,
        source_ref=row.get("source_ref") or template_name,
    )


def _flat_row_to_components(raw_row: dict) -> list[dict]:
    """Convert flat CSV columns into the Karix components array."""
    components = []

    # 1. HEADER
    htype = (raw_row.get("header_type") or raw_row.get("header_format") or "").strip().upper()
    if htype in ("IMAGE", "VIDEO", "DOCUMENT"):
        comp = {"type": "HEADER", "format": htype}
        if raw_row.get("header_media_url") or raw_row.get("media_url"):
            comp["media_url"] = (raw_row.get("header_media_url") or raw_row.get("media_url")).strip()
        elif raw_row.get("header_media_file") or raw_row.get("media_file"):
            comp["media_file"] = (raw_row.get("header_media_file") or raw_row.get("media_file")).strip()
        components.append(comp)
    elif htype == "LOCATION":
        components.append({"type": "HEADER", "format": "LOCATION"})
    elif (htype in ("TEXT", "HEADER") or (not htype and (raw_row.get("header_text") or raw_row.get("header")))) and (
        raw_row.get("header_text") or raw_row.get("header")
    ):
        components.append(
            {
                "type": "HEADER",
                "format": "TEXT",
                "text": (raw_row.get("header_text") or raw_row.get("header")).strip(),
            }
        )
    # 2. BODY
    body_text = (raw_row.get("body") or raw_row.get("body_text") or "").strip()
    if body_text:
        # Normalize line endings and collapse 3+ consecutive newlines (Meta Rule: max 2)
        body_text = body_text.replace("\r\n", "\n").replace("\r", "\n")
        body_text = re.sub(r"[ \t]+\n", "\n", body_text)
        body_text = re.sub(r"\n{3,}", "\n\n", body_text)
        body_text = re.sub(r"[ \t]+([.,!?:;])", r"\1", body_text)
        body_text = re.sub(r"([A-Za-z0-9])(<[^>]+>)", r"\1 \2", body_text)
        body_text = re.sub(r"(<[^>]+>)([A-Za-z0-9])", r"\1 \2", body_text)
        body_text = re.sub(r"([A-Za-z0-9])(\{#[^#]+#\})", r"\1 \2", body_text)
        placeholders = []

        def _repl(m):
            idx = len(placeholders) + 1
            placeholders.append(m.group(0))
            return f"{{{{{idx}}}}}"

        pattern = r"(\{\{\d+\}\}|\{\{[a-zA-Z0-9_]+\}\}|<[^>]+>|\{#[^#]+#\}|\[[a-zA-Z0-9_]+\]|\{[a-zA-Z0-9_]+\})"
        body_text = re.sub(pattern, _repl, body_text)
        components.append({"type": "BODY", "text": body_text.strip()})
    # 3. FOOTER
    footer_text = (raw_row.get("footer") or raw_row.get("footer_text") or "").strip()
    if footer_text:
        components.append({"type": "FOOTER", "text": footer_text})

    # 4. BUTTONS
    btype = (raw_row.get("button_type") or "").strip().upper()
    btext = (raw_row.get("button_text") or "").strip()
    burl = (raw_row.get("button_url") or "").strip()
    bexample = (raw_row.get("button_url_example") or "").strip() or "https://www.tatacapital.com/personal-loan.html"

    if btype in ("URL", "") and btext and burl:
        btn_data = {
            "type": "URL",
            "text": btext,
            "url": burl,
        }
        # Meta Rule: 'example' is ONLY valid if URL contains dynamic variables (e.g. {{1}})
        if "{{1}}" in burl or "{{0}}" in burl or "<" in burl:
            btn_data["example"] = [bexample]
        components.append(
            {
                "type": "BUTTONS",
                "buttons": [btn_data],
            }
        )
    elif btype in ("QUICK_REPLY", "QUICKREPLY") and btext:
        btn_items = [{"type": "QUICK_REPLY", "text": item.strip()} for item in btext.split("|") if item.strip()]
        if btn_items:
            components.append(
                {
                    "type": "BUTTONS",
                    "buttons": btn_items,
                }
            )

    return components


def _resolve_row_waba(row_client: str, cache: dict[str, str]) -> str:
    """Resolve the WABA ID for a row's client, cached per client per file load."""
    c = (row_client or "bajaj").lower()
    if c not in cache:
        try:
            cache[c] = get_waba_id(c) or ""
        except Exception:
            cache[c] = ""
    return cache[c]


def load_from_csv(path: str, client: str = "bajaj") -> list[TemplateSubmission]:
    """Load templates from a CSV file for the specified client."""
    waba_cache: dict[str, str] = {}
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for raw_row in csv.DictReader(f):
            clean_row = {k.strip(): (v.strip() if v else "") for k, v in raw_row.items() if k}
            if not clean_row.get("template_name") and not clean_row.get("name"):
                continue

            if clean_row.get("components"):
                try:
                    clean_row["components"] = json.loads(clean_row["components"])
                except json.JSONDecodeError:
                    clean_row["components"] = _flat_row_to_components(clean_row)
            else:
                clean_row["components"] = _flat_row_to_components(clean_row)

            clean_row["client"] = clean_row.get("client") or client
            if not clean_row.get("waba_id"):
                clean_row["waba_id"] = _resolve_row_waba(clean_row["client"], waba_cache)
            if not clean_row.get("language"):
                clean_row["language"] = "en"
            if not clean_row.get("category"):
                clean_row["category"] = "MARKETING"

            rows.append(clean_row)

    # Only keep rows that actually produced template content. A non-template
    # file (e.g. a customer-data export with a `name` column) would otherwise
    # be turned into empty, invalid templates instead of being ignored.
    submissions = [_row_to_submission(row, client=client) for row in rows]
    kept = [s for s in submissions if s.components]
    for dropped in (s for s in submissions if not s.components):
        logger.warning(
            "Dropping row %r from %s: no template components parsed",
            dropped.template_name,
            path,
        )
    return kept


def _build_single_cell_card_submission(
    idx: int, block: str, client: str, clean_name: str, extracted_media: list, waba_cache: dict
) -> TemplateSubmission:
    parsed = parse_single_cell_whatsapp_block(block, client=client)
    t_name = f"{client.lower()}_{clean_name}_card_{idx}"[:30]

    components = []
    if (idx - 1) < len(extracted_media):
        m = extracted_media[idx - 1]
        components.append(
            {
                "type": "HEADER",
                "format": m["kind"],
                "image_bytes": m["bytes"],
                "file_type": m["mime_type"],
            }
        )
    elif parsed.get("header"):
        components.append(
            {
                "type": "HEADER",
                "format": "TEXT",
                "text": parsed["header"],
            }
        )
    components.append({"type": "BODY", "text": parsed["body"]})
    components.append({"type": "FOOTER", "text": "T&Cs apply"})
    components.append(
        {
            "type": "BUTTONS",
            "buttons": [
                {
                    "type": "URL",
                    "text": parsed["button_text"],
                    "url": parsed["button_url"],
                    "example": [parsed["button_example"]],
                }
            ],
        }
    )

    return TemplateSubmission(
        client=client.lower(),
        channel="whatsapp",
        template_name=t_name,
        language="en",
        category="MARKETING",
        waba_id=_resolve_row_waba(client, waba_cache),
        components=components,
        source_ref=t_name,
    )


def _parse_single_cell_excel_blocks(
    all_raw_rows: list, path: str, client: str, extracted_media: list, waba_cache: dict
) -> list[TemplateSubmission] | None:
    block_row_cells = None
    for r in all_raw_rows:
        text_blocks = [str(c).strip() for c in r if c is not None and len(str(c).strip()) > 35]
        if len(text_blocks) >= 2:
            block_row_cells = text_blocks
            break
    if not block_row_cells:
        return None

    base_name = Path(path).stem
    clean_name = re.sub(r"[^a-zA-Z0-9_]", "_", base_name).strip("_").lower()

    return [
        _build_single_cell_card_submission(idx, block, client, clean_name, extracted_media, waba_cache)
        for idx, block in enumerate(block_row_cells, 1)
    ]


def load_from_excel(path: str, client: str = "bajaj") -> list[TemplateSubmission]:
    """
    Load templates from an Excel (.xlsx) file.
    Supports embedded images/videos/documents, multi-block cards, and standard column tables.
    """
    import openpyxl

    waba_cache: dict[str, str] = {}
    extracted_media = _extract_media_from_xlsx(path)
    videos = [m for m in extracted_media if m["kind"] == "VIDEO"]
    images = [m for m in extracted_media if m["kind"] == "IMAGE"]
    docs = [m for m in extracted_media if m["kind"] == "DOCUMENT"]
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    all_raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    first_row = [str(c or "").strip().lower() for c in all_raw_rows[0]] if all_raw_rows else []
    has_standard_headers = any(
        h
        in (
            "template_name",
            "name",
            "body",
            "body_text",
            "components",
            "category",
            "language",
        )
        for h in first_row
    )
    block_subs = _parse_single_cell_excel_blocks(all_raw_rows, path, client, extracted_media, waba_cache) if not has_standard_headers else None
    if block_subs is not None:
        return block_subs

    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = []
    v_idx = 0
    img_idx = 0
    doc_idx = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        raw_row = {}
        for h, val in zip(headers, row, strict=False):
            if h:
                raw_row[h] = str(val).strip() if val is not None else ""

        if not raw_row.get("template_name") and not raw_row.get("name"):
            continue

        if raw_row.get("components"):
            try:
                raw_row["components"] = json.loads(raw_row["components"])
            except json.JSONDecodeError:
                raw_row["components"] = _flat_row_to_components(raw_row)
        else:
            raw_row["components"] = _flat_row_to_components(raw_row)

        v_idx, img_idx, doc_idx = _bind_embedded_media_to_row(
            raw_row, videos, images, docs, extracted_media, v_idx, img_idx, doc_idx
        )
        raw_row["client"] = raw_row.get("client") or client
        if not raw_row.get("waba_id"):
            raw_row["waba_id"] = _resolve_row_waba(raw_row["client"], waba_cache)
        if not raw_row.get("language"):
            raw_row["language"] = "en"
        if not raw_row.get("category"):
            raw_row["category"] = "MARKETING"

        rows.append(raw_row)

    submissions = [_row_to_submission(row, client=client) for row in rows]
    kept = [s for s in submissions if s.components]
    for dropped in (s for s in submissions if not s.components):
        logger.warning(
            "Dropping row %r from %s: no template components parsed",
            dropped.template_name,
            path,
        )
    return kept


def load_from_json(path: str) -> list[TemplateSubmission]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return [_row_to_submission(row) for row in data]


def load_from_list(rows: list[dict], client: str = "bajaj") -> list[TemplateSubmission]:
    return [_row_to_submission(row, client=client) for row in rows]
def _bind_embedded_media_to_row(
    raw_row: dict,
    videos: list,
    images: list,
    docs: list,
    extracted_media: list,
    v_idx: int,
    img_idx: int,
    doc_idx: int,
) -> tuple[int, int, int]:
    for comp in raw_row.get("components", []):
        if not isinstance(comp, dict) or comp.get("type") != "HEADER":
            continue
        cformat = str(comp.get("format", "")).upper()
        if comp.get("media_url") or comp.get("media_file"):
            continue
        if cformat == "VIDEO":
            if v_idx < len(videos):
                comp["image_bytes"] = videos[v_idx]["bytes"]
                comp["file_type"] = videos[v_idx]["mime_type"]
                v_idx += 1
            elif v_idx < len(extracted_media):
                comp["image_bytes"] = extracted_media[v_idx]["bytes"]
                comp["file_type"] = "video/mp4"
                v_idx += 1
        elif cformat == "IMAGE":
            if img_idx < len(images):
                comp["image_bytes"] = images[img_idx]["bytes"]
                comp["file_type"] = images[img_idx]["mime_type"]
                img_idx += 1
            elif img_idx < len(extracted_media):
                comp["image_bytes"] = extracted_media[img_idx]["bytes"]
                comp["file_type"] = "image/png"
                img_idx += 1
        elif cformat == "DOCUMENT":
            if doc_idx < len(docs):
                comp["image_bytes"] = docs[doc_idx]["bytes"]
                comp["file_type"] = docs[doc_idx]["mime_type"]
                doc_idx += 1
            elif doc_idx < len(extracted_media):
                comp["image_bytes"] = extracted_media[doc_idx]["bytes"]
                comp["file_type"] = "application/pdf"
                doc_idx += 1
    return v_idx, img_idx, doc_idx


def _score_single_row_routing(
    explicit_client: str, raw_header_val: str, comp_headers: list[str], comp_text: str, tname: str, scores: dict, reasons: dict
) -> None:
    for rule in ACCOUNT_ROUTING_RULES:
        r_id = rule["id"]
        if explicit_client and r_id in explicit_client:
            scores[r_id] += 20
            if f"Explicit account column: '{explicit_client}'" not in reasons[r_id]:
                reasons[r_id].append(f"Explicit account column: '{explicit_client}'")
        for tag in rule["headers"]:
            if tag in raw_header_val or any(tag in ch for ch in comp_headers):
                scores[r_id] += 10
                if f"Header tag '{tag}'" not in reasons[r_id]:
                    reasons[r_id].append(f"Header tag '{tag}'")
        for prefix in rule["name_prefixes"]:
            if tname.startswith(prefix):
                scores[r_id] += 5
                if f"Template prefix '{prefix}'" not in reasons[r_id]:
                    reasons[r_id].append(f"Template prefix '{prefix}'")
        for kw in rule["keywords"]:
            if kw in comp_text or kw in tname:
                scores[r_id] += 2
                if f"Keyword '{kw}'" not in reasons[r_id]:
                    reasons[r_id].append(f"Keyword '{kw}'")




ACCOUNT_ROUTING_RULES = [
    {
        "id": "tchfl",
        "name": "TCHFL — Housing Finance (HLTATA)",
        "headers": ["HLTATA", "TCHFL", "HL_TATA", "HOUSING"],
        "name_prefixes": ["hfl_", "tchfl_", "hl_"],
        "keywords": ["housing finance", "home loan", "tchfl", "hltata", "housing"],
    },
    {
        "id": "tcl_promo",
        "name": "TCL — Promotional (PL, BL, UCL, NCL, LAP)",
        "headers": ["PLTATA", "TATABL", "ALTATA", "TCLLAP", "Tatacl", "TCLPROMO"],
        "name_prefixes": ["pl_", "bl_", "ucl_", "ncl_", "lap_", "promo_"],
        "keywords": ["personal loan", "business loan", "used car loan", "two wheeler", "promotional", "pltata", "tatabl"],
    },
    {
        "id": "tcl_trans",
        "name": "TCL — Transactional (LAP, Services & Collections)",
        "headers": ["TCLLAP", "Tatacl", "TCLTRANS"],
        "name_prefixes": ["trans_", "srv_", "coll_", "serv_"],
        "keywords": ["transactional", "emi reminder", "overdue", "collections", "statement"],
    },
    {
        "id": "wealth",
        "name": "Tata Wealth & Securities (TATAWL)",
        "headers": ["TATAWL", "WEALTH", "TATA_WEALTH"],
        "name_prefixes": ["wl_", "wealth_", "sec_"],
        "keywords": ["wealth", "securities", "portfolio", "demat", "trading", "tatawl"],
    },
    {
        "id": "moneyfy",
        "name": "Tata Moneyfy",
        "headers": ["MONEYFY", "TATA_MONEYFY"],
        "name_prefixes": ["mf_", "moneyfy_"],
        "keywords": ["moneyfy", "mutual fund", "sip", "investment"],
    },
    {
        "id": "bajaj",
        "name": "Bajaj Finserv",
        "headers": ["BAJAJ", "BFDL", "BAJAJ_FINSERV"],
        "name_prefixes": ["bajaj_", "bfdl_"],
        "keywords": ["bajaj", "finserv", "bfdl", "emi card", "bfl"],
    },
]


def detect_spreadsheet_account(
    submissions_or_rows: list,
    current_account: str = "bajaj",
) -> dict:
    """
    Inspect spreadsheet rows, template names, header tags, and body text
    to infer the target sub-account with a confidence score and explanation.
    """
    if not submissions_or_rows:
        return {
            "detected_account_id": current_account,
            "detected_account_name": current_account.title(),
            "confidence": 0.0,
            "matched_reasons": ["No rows to inspect"],
            "is_mismatch": False,
            "current_account": current_account,
        }

    scores = {rule["id"]: 0 for rule in ACCOUNT_ROUTING_RULES}
    reasons = {rule["id"]: [] for rule in ACCOUNT_ROUTING_RULES}

    for item in submissions_or_rows:
        row = item if isinstance(item, dict) else (getattr(item, "__dict__", None) or {})
        tname = str(row.get("template_name") or getattr(item, "template_name", "") or "").lower()

        # Check components text
        comp_text = ""
        comp_headers = []
        components = row.get("components") or getattr(item, "components", []) or []
        for comp in components:
            c = comp if isinstance(comp, dict) else (getattr(comp, "__dict__", None) or {})
            if c.get("text"):
                comp_text += " " + str(c.get("text", "")).lower()
            if c.get("type") == "HEADER":
                h_text = str(c.get("text") or "").upper()
                if h_text:
                    comp_headers.append(h_text)

        # Check raw header columns (e.g. sender_id, header, client, sub_product)
        explicit_client = str(row.get("client") or row.get("account") or row.get("sub_product") or "").lower()
        raw_header_val = str(row.get("header") or row.get("sender_id") or row.get("dlt_header") or "").upper()

        _score_single_row_routing(explicit_client, raw_header_val, comp_headers, comp_text, tname, scores, reasons)
    # Pick highest score
    sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    best_id, best_score = sorted_scores[0]

    # Calculate confidence (0.0 to 1.0)
    total_score = sum(scores.values())
    confidence = (best_score / total_score) if total_score > 0 else 0.0

    if best_score == 0:
        best_id = current_account
        confidence = 0.0
        reasons[best_id] = ["No distinct brand indicators found; retained current account."]

    rule_meta = next((r for r in ACCOUNT_ROUTING_RULES if r["id"] == best_id), None)
    best_name = rule_meta["name"] if rule_meta else best_id.title()

    curr_norm = (current_account or "bajaj").lower().strip()
    is_mismatch = (best_id != curr_norm) and (confidence >= 0.35)

    return {
        "detected_account_id": best_id,
        "detected_account_name": best_name,
        "confidence": round(confidence, 2),
        "matched_reasons": reasons.get(best_id, [])[:3],
        "is_mismatch": is_mismatch,
        "current_account": current_account,
        "all_scores": {k: v for k, v in scores.items() if v > 0},
    }
